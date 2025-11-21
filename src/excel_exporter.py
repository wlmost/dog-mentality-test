"""
Excel-Export für Test-Sessions

Exportiert Stammdaten und Testergebnisse in eine strukturierte Excel-Datei
"""
from pathlib import Path
from typing import Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from src.test_session import TestSession
from src.test_battery import TestBattery


class ExcelExportError(Exception):
    """Fehler beim Excel-Export"""
    pass


class ExcelExporter:
    """
    Exportiert Test-Sessions in Excel-Format
    
    Features:
    - Zwei Sheets: Stammdaten und Testergebnisse
    - Formatierung: Header fett, gefärbt
    - Auto-Spaltenbreite
    """
    
    def __init__(self, battery: Optional[TestBattery] = None):
        """
        Initialisiert den Exporter
        
        Args:
            battery: Testbatterie für zusätzliche Test-Informationen (optional)
        """
        self._battery = battery
    
    def export_to_excel(self, session: TestSession, filepath: str):
        """
        Exportiert Session in Excel-Datei
        
        Args:
            session: Test-Session mit Daten
            filepath: Pfad zur Ausgabedatei (.xlsx)
            
        Raises:
            ExcelExportError: Bei Fehlern während des Exports
        """
        try:
            # Workbook erstellen
            wb = Workbook()
            
            # Sheet 1: Stammdaten
            ws_master = wb.active
            ws_master.title = "Stammdaten"
            self._write_master_data(ws_master, session)
            
            # Sheet 2: Testergebnisse
            ws_results = wb.create_sheet("Testergebnisse")
            self._write_test_results(ws_results, session)
            
            # Sheet 3: OCEAN-Analyse (nur wenn Profile vorhanden)
            if session.ideal_profile or session.owner_profile or (self._battery and session.results):
                ws_ocean = wb.create_sheet("OCEAN-Analyse")
                self._write_ocean_profiles(ws_ocean, session)
            
            # Sheet 4: KI-Bewertung (nur wenn vorhanden)
            if session.ai_assessment:
                ws_ai = wb.create_sheet("KI-Bewertung")
                self._write_ai_assessment(ws_ai, session)
            
            # Speichern
            wb.save(filepath)
            
        except PermissionError as e:
            raise ExcelExportError(
                f"Datei ist möglicherweise geöffnet oder schreibgeschützt: {filepath}"
            ) from e
        except Exception as e:
            raise ExcelExportError(
                f"Fehler beim Excel-Export: {str(e)}"
            ) from e
    
    def _write_master_data(self, ws, session: TestSession):
        """Schreibt Stammdaten in Worksheet"""
        # Header-Style
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        
        # Titel
        ws["A1"] = "Stammdaten"
        ws["A1"].font = Font(bold=True, size=14)
        
        # Überschriften
        headers = ["Feld", "Wert"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Datum formatieren (session.date ist ISO-String)
        try:
            date_obj = datetime.fromisoformat(session.date)
            date_str = date_obj.strftime("%d.%m.%Y %H:%M")
        except (ValueError, AttributeError):
            date_str = session.date
        
        # Daten
        data = [
            ("Datum", date_str),
            ("Name des Halters", session.dog_data.owner_name),
            ("Name des Hundes", session.dog_data.dog_name),
            ("Rasse", session.dog_data.breed if session.dog_data.breed else "-"),
            ("Alter", session.dog_data.age_display()),
            ("Geschlecht", session.dog_data.gender.value),
            ("Kastriert", "Ja" if session.dog_data.neutered else "Nein"),
            ("Zukünftiges Einsatzgebiet", session.dog_data.intended_use if session.dog_data.intended_use else "-"),
            ("Testbatterie", session.battery_name),
        ]
        
        for row, (field, value) in enumerate(data, start=4):
            ws.cell(row=row, column=1, value=field)
            ws.cell(row=row, column=2, value=value)
        
        # Session-Notizen
        if session.session_notes:
            ws.cell(row=len(data) + 5, column=1, value="Session-Notizen")
            ws.cell(row=len(data) + 5, column=1).font = Font(bold=True)
            ws.cell(row=len(data) + 6, column=1, value=session.session_notes)
            ws.merge_cells(f"A{len(data) + 6}:B{len(data) + 6}")
        
        # Spaltenbreiten anpassen
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 40
    
    def _write_test_results(self, ws, session: TestSession):
        """Schreibt Testergebnisse in Worksheet"""
        # Header-Style
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        
        # Titel
        ws["A1"] = "Testergebnisse"
        ws["A1"].font = Font(bold=True, size=14)
        
        # Überschriften
        headers = ["Nr.", "Testname", "OCEAN-Dimension", "Score", "Notizen"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Daten aus Session
        row = 4
        for test_number, result in sorted(session.results.items()):
            ws.cell(row=row, column=1, value=test_number)
            
            # Testname und OCEAN-Dimension aus Battery (falls vorhanden)
            if self._battery:
                test = self._battery.get_test_by_number(test_number)
                if test:
                    ws.cell(row=row, column=2, value=test.name)
                    ws.cell(row=row, column=3, value=test.ocean_dimension.value)
                else:
                    ws.cell(row=row, column=2, value="—")
                    ws.cell(row=row, column=3, value="—")
            else:
                ws.cell(row=row, column=2, value="—")
                ws.cell(row=row, column=3, value="—")
            
            ws.cell(row=row, column=4, value=result.score)
            ws.cell(row=row, column=5, value=result.notes or "")
            
            row += 1
        
        # Spaltenbreiten anpassen
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 50
        
        # Score-Spalte zentrieren
        for row_idx in range(4, row):
            ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal="center")
    
    def _write_ocean_profiles(self, ws, session: TestSession):
        """Schreibt OCEAN-Profile in Worksheet (Ist, Fragebogen, Ideal)"""
        from src.ocean_analyzer import OceanAnalyzer
        
        # Header-Style
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        
        # Titel
        ws["A1"] = "OCEAN-Persönlichkeitsanalyse"
        ws["A1"].font = Font(bold=True, size=14)
        
        # Überschriften
        headers = ["Dimension", "Ist-Profil", "Fragebogen-Profil", "Ideal-Profil"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # OCEAN-Dimensionen berechnen (Ist-Profil)
        ocean_scores = None
        if self._battery and session.results:
            analyzer = OceanAnalyzer(session, self._battery)
            ocean_scores = analyzer.calculate_ocean_scores()
        
        # Dimensionsnamen
        dimensions = [
            ("Offenheit (O)", "O", "openness"),
            ("Gewissenhaftigkeit (C)", "C", "conscientiousness"),
            ("Extraversion (E)", "E", "extraversion"),
            ("Verträglichkeit (A)", "A", "agreeableness"),
            ("Neurotizismus (N)", "N", "neuroticism")
        ]
        
        # Daten schreiben
        for row, (dim_name, dim_key, dim_attr) in enumerate(dimensions, start=4):
            # Spalte 1: Dimensionsname
            ws.cell(row=row, column=1, value=dim_name)
            ws.cell(row=row, column=1).font = Font(bold=True)
            
            # Spalte 2: Ist-Profil (aus Testergeb nissen berechnet)
            if ocean_scores:
                ist_value = getattr(ocean_scores, dim_attr, 0)
                ws.cell(row=row, column=2, value=ist_value)
                ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
            else:
                ws.cell(row=row, column=2, value="—")
                ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
            
            # Spalte 3: Fragebogen-Profil (Halter-Erwartungen)
            if session.owner_profile and dim_key in session.owner_profile:
                ws.cell(row=row, column=3, value=session.owner_profile[dim_key])
                ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
            else:
                ws.cell(row=row, column=3, value="—")
                ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
            
            # Spalte 4: Ideal-Profil (KI-generiert)
            if session.ideal_profile and dim_key in session.ideal_profile:
                ws.cell(row=row, column=4, value=session.ideal_profile[dim_key])
                ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
            else:
                ws.cell(row=row, column=4, value="—")
                ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
        
        # Spaltenbreiten anpassen
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 15
    
    def _write_ai_assessment(self, ws, session: TestSession):
        """Schreibt KI-Bewertung in Worksheet"""
        # Titel
        ws["A1"] = "KI-Bewertung"
        ws["A1"].font = Font(bold=True, size=14)
        
        # Info-Text
        ws["A3"] = "Diese Bewertung wurde von einer KI basierend auf dem Ist-Profil, Fragebogen-Profil und Ideal-Profil erstellt."
        ws["A3"].font = Font(size=10, italic=True)
        ws.merge_cells("A3:D3")
        
        # Bewertungstext
        ws["A5"] = session.ai_assessment
        ws["A5"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells("A5:D20")
        
        # Spaltenbreiten
        ws.column_dimensions["A"].width = 100
        ws.row_dimensions[5].height = 300
