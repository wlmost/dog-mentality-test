"""
Excel-Parser für Testbatterien
"""
import openpyxl
from pathlib import Path
from typing import Optional
from src.test_battery import Test, TestBattery, OceanDimension


class ExcelImportError(Exception):
    """Fehler beim Excel-Import"""
    pass


class TestBatteryImporter:
    """
    Importiert Testbatterie aus Excel-Datei
    """
    
    # Spalten-Indizes (0-basiert)
    COL_NUMBER = 0
    COL_OCEAN = 1
    COL_NAME = 2
    COL_SETTING = 3
    COL_MATERIALS = 4
    COL_DURATION = 5
    COL_FIGURANT = 6
    COL_CRITERIA = 7
    COL_SCALE = 8
    
    def __init__(self, filepath: str):
        """
        Initialisiert den Importer
        
        Args:
            filepath: Pfad zur Excel-Datei
        """
        self.filepath = Path(filepath)
        
        if not self.filepath.suffix.lower() in ['.xlsx', '.xls']:
            raise ValueError(f"Ungültiges Dateiformat: {self.filepath.suffix}")
        
        if not self.filepath.exists():
            raise FileNotFoundError(f"Datei nicht gefunden: {filepath}")
    
    def import_battery(self, sheet_name: Optional[str] = None) -> TestBattery:
        """
        Importiert Testbatterie aus Excel
        
        Args:
            sheet_name: Name des Worksheets (None = aktives Sheet)
            
        Returns:
            TestBattery-Objekt
            
        Raises:
            ExcelImportError: Bei Fehlern beim Import
        """
        try:
            workbook = openpyxl.load_workbook(self.filepath, data_only=True)
            
            # Worksheet auswählen
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    raise ExcelImportError(
                        f"Sheet '{sheet_name}' nicht gefunden. "
                        f"Verfügbare Sheets: {workbook.sheetnames}"
                    )
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active
            
            # Tests einlesen (Zeile 2 bis Ende, Zeile 1 ist Header)
            tests = []
            battery_name = sheet_name or worksheet.title
            
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                # Leere Zeilen überspringen
                if not row[self.COL_NUMBER]:
                    continue
                
                try:
                    test = self._parse_test_row(row, row_idx)
                    tests.append(test)
                except Exception as e:
                    raise ExcelImportError(
                        f"Fehler in Zeile {row_idx}: {str(e)}"
                    )
            
            if not tests:
                raise ExcelImportError("Keine Tests gefunden")
            
            return TestBattery(name=battery_name, tests=tests)
            
        except Exception as e:
            if isinstance(e, ExcelImportError):
                raise
            raise ExcelImportError(f"Unerwarteter Fehler beim Import: {str(e)}")
    
    def _parse_test_row(self, row: tuple, row_idx: int) -> Test:
        """
        Parst eine Zeile zu einem Test-Objekt
        
        Args:
            row: Zeile als Tuple
            row_idx: Zeilennummer (für Fehlermeldungen)
            
        Returns:
            Test-Objekt
        """
        # Test-Nummer
        try:
            number = int(row[self.COL_NUMBER])
        except (ValueError, TypeError):
            raise ValueError(f"Ungültige Test-Nummer: {row[self.COL_NUMBER]}")
        
        # OCEAN-Dimension
        ocean_text = str(row[self.COL_OCEAN] or "").strip()
        ocean_dim = None
        for dim in OceanDimension:
            if dim.value == ocean_text:
                ocean_dim = dim
                break
        
        if ocean_dim is None:
            raise ValueError(f"Unbekannte OCEAN-Dimension: {ocean_text}")
        
        # Strings mit Default-Werten
        name = str(row[self.COL_NAME] or "").strip()
        setting = str(row[self.COL_SETTING] or "").strip()
        materials = str(row[self.COL_MATERIALS] or "").strip()
        duration = str(row[self.COL_DURATION] or "").strip()
        figurant = str(row[self.COL_FIGURANT] or "").strip()
        criteria = str(row[self.COL_CRITERIA] or "").strip()
        scale = str(row[self.COL_SCALE] or "").strip()
        
        return Test(
            number=number,
            ocean_dimension=ocean_dim,
            name=name,
            setting=setting,
            materials=materials,
            duration=duration,
            role_figurant=figurant,
            observation_criteria=criteria,
            rating_scale=scale
        )
    
    def get_sheet_names(self) -> list[str]:
        """Gibt Liste der verfügbaren Sheet-Namen zurück"""
        workbook = openpyxl.load_workbook(self.filepath, read_only=True)
        return workbook.sheetnames
