"""
Tests für Excel-Exporter
"""
import pytest
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

from src.excel_exporter import ExcelExporter, ExcelExportError
from src.test_session import TestSession, TestResult
from src.models import DogData, Gender
from src.test_battery import TestBattery, Test, OceanDimension


@pytest.fixture
def sample_session():
    """Erstellt eine Test-Session mit Daten"""
    dog_data = DogData(
        owner_name="Max Mustermann",
        dog_name="Bello",
        age_years=3,
        age_months=6,
        gender=Gender.MALE,
        neutered=True
    )
    
    session = TestSession(
        dog_data=dog_data,
        battery_name="OCEAN Testbatterie"
    )
    
    # Einige Testergebnisse hinzufügen
    session.add_result(TestResult(test_number=1, score=2, notes="Sehr aufgeschlossen"))
    session.add_result(TestResult(test_number=2, score=-1, notes="Etwas zurückhaltend"))
    session.add_result(TestResult(test_number=5, score=0, notes=""))
    
    session.session_notes = "Hund war heute etwas müde, aber kooperativ."
    
    return session


@pytest.fixture
def sample_battery():
    """Erstellt eine Test-Batterie"""
    tests = [
        Test(
            number=1,
            ocean_dimension=OceanDimension.OPENNESS,
            name="Neugierverhalten",
            setting="Innenraum",
            materials="Spielzeug",
            duration="5 min",
            role_figurant="Beobachter",
            observation_criteria="Interesse",
            rating_scale="-2 bis +2"
        ),
        Test(
            number=2,
            ocean_dimension=OceanDimension.AGREEABLENESS,
            name="Sozialverhalten",
            setting="Außenbereich",
            materials="Keine",
            duration="10 min",
            role_figurant="Aktiv",
            observation_criteria="Freundlichkeit",
            rating_scale="-2 bis +2"
        ),
        Test(
            number=5,
            ocean_dimension=OceanDimension.EXTRAVERSION,
            name="Aktivitätslevel",
            setting="Freigelände",
            materials="Ball",
            duration="15 min",
            role_figurant="Spielpartner",
            observation_criteria="Energie",
            rating_scale="-2 bis +2"
        ),
    ]
    
    return TestBattery(name="OCEAN Testbatterie", tests=tests)


class TestExcelExporter:
    """Tests für ExcelExporter"""
    
    def test_exporter_creation(self):
        """Test: Exporter kann erstellt werden"""
        exporter = ExcelExporter()
        assert exporter is not None
    
    def test_exporter_with_battery(self, sample_battery):
        """Test: Exporter kann mit Battery erstellt werden"""
        exporter = ExcelExporter(battery=sample_battery)
        assert exporter is not None
    
    def test_export_creates_file(self, sample_session, tmp_path):
        """Test: Export erstellt Excel-Datei"""
        exporter = ExcelExporter()
        filepath = tmp_path / "test_export.xlsx"
        
        exporter.export_to_excel(sample_session, str(filepath))
        
        assert filepath.exists()
    
    def test_export_has_two_sheets(self, sample_session, tmp_path):
        """Test: Excel hat zwei Sheets (Stammdaten und Testergebnisse)"""
        exporter = ExcelExporter()
        filepath = tmp_path / "test_export.xlsx"
        
        exporter.export_to_excel(sample_session, str(filepath))
        
        wb = load_workbook(str(filepath))
        assert "Stammdaten" in wb.sheetnames
        assert "Testergebnisse" in wb.sheetnames
    
    def test_master_data_sheet_content(self, sample_session, tmp_path):
        """Test: Stammdaten-Sheet enthält korrekte Daten"""
        exporter = ExcelExporter()
        filepath = tmp_path / "test_export.xlsx"
        
        exporter.export_to_excel(sample_session, str(filepath))
        
        wb = load_workbook(str(filepath))
        ws = wb["Stammdaten"]
        
        # Prüfe einige Zellen
        assert ws["A1"].value == "Stammdaten"
        
        # Finde Zeilen mit bestimmten Daten
        cells_text = []
        for row in ws.iter_rows(min_row=4, max_row=15, min_col=1, max_col=2):
            cells_text.append((row[0].value, row[1].value))
        
        # Prüfe ob wichtige Daten vorhanden sind
        cell_dict = dict(cells_text)
        assert cell_dict.get("Name des Halters") == "Max Mustermann"
        assert cell_dict.get("Name des Hundes") == "Bello"
        assert cell_dict.get("Geschlecht") == "Rüde"
    
    def test_results_sheet_content(self, sample_session, tmp_path):
        """Test: Testergebnisse-Sheet enthält korrekte Daten"""
        exporter = ExcelExporter()
        filepath = tmp_path / "test_export.xlsx"
        
        exporter.export_to_excel(sample_session, str(filepath))
        
        wb = load_workbook(str(filepath))
        ws = wb["Testergebnisse"]
        
        # Prüfe Titel
        assert ws["A1"].value == "Testergebnisse"
        
        # Prüfe Header
        assert ws["A3"].value == "Nr."
        assert ws["B3"].value == "Testname"
        assert ws["C3"].value == "OCEAN-Dimension"
        assert ws["D3"].value == "Score"
        assert ws["E3"].value == "Notizen"
        
        # Prüfe erste Datenzeile
        assert ws["A4"].value == 1
        assert ws["D4"].value == 2
        assert ws["E4"].value == "Sehr aufgeschlossen"
    
    def test_export_with_battery_includes_test_names(self, sample_session, sample_battery, tmp_path):
        """Test: Export mit Battery enthält Testnamen"""
        exporter = ExcelExporter(battery=sample_battery)
        filepath = tmp_path / "test_export.xlsx"
        
        exporter.export_to_excel(sample_session, str(filepath))
        
        wb = load_workbook(str(filepath))
        ws = wb["Testergebnisse"]
        
        # Erste Zeile sollte Testname haben
        assert ws["B4"].value == "Neugierverhalten"
        assert ws["C4"].value == "Offenheit"
    
    def test_export_without_battery_has_placeholders(self, sample_session, tmp_path):
        """Test: Export ohne Battery hat Platzhalter für Testnamen"""
        exporter = ExcelExporter()
        filepath = tmp_path / "test_export.xlsx"
        
        exporter.export_to_excel(sample_session, str(filepath))
        
        wb = load_workbook(str(filepath))
        ws = wb["Testergebnisse"]
        
        # Sollte Platzhalter haben
        assert ws["B4"].value == "—"
        assert ws["C4"].value == "—"
    
    def test_export_includes_session_notes(self, sample_session, tmp_path):
        """Test: Session-Notizen werden exportiert"""
        exporter = ExcelExporter()
        filepath = tmp_path / "test_export.xlsx"
        
        exporter.export_to_excel(sample_session, str(filepath))
        
        wb = load_workbook(str(filepath))
        ws = wb["Stammdaten"]
        
        # Finde Session-Notizen
        found_notes = False
        for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=2):
            if row[0].value and "Hund war heute" in str(row[0].value):
                found_notes = True
                break
        
        assert found_notes
    
    def test_export_handles_permission_error(self, sample_session, tmp_path, monkeypatch):
        """Test: Fehlerbehandlung bei Schreibfehler"""
        exporter = ExcelExporter()
        filepath = tmp_path / "test_export.xlsx"
        
        # Mock Workbook.save um PermissionError zu simulieren
        def mock_save(*args, **kwargs):
            raise PermissionError("File is locked")
        
        from openpyxl import Workbook
        monkeypatch.setattr(Workbook, "save", mock_save)
        
        with pytest.raises(ExcelExportError) as exc_info:
            exporter.export_to_excel(sample_session, str(filepath))
        
        assert "schreibgeschützt" in str(exc_info.value).lower()
    
    def test_export_multiple_results(self, sample_battery, tmp_path):
        """Test: Export mit vielen Testergebnissen"""
        dog_data = DogData(
            owner_name="Test Owner",
            dog_name="Test Dog",
            age_years=2,
            age_months=0,
            gender=Gender.FEMALE,
            neutered=False
        )
        
        session = TestSession(dog_data=dog_data, battery_name="Test Battery")
        
        # Viele Ergebnisse hinzufügen
        for i in range(1, 11):
            session.add_result(TestResult(test_number=i, score=i % 5 - 2, notes=f"Note {i}"))
        
        exporter = ExcelExporter(battery=sample_battery)
        filepath = tmp_path / "test_export.xlsx"
        
        exporter.export_to_excel(session, str(filepath))
        
        wb = load_workbook(str(filepath))
        ws = wb["Testergebnisse"]
        
        # Prüfe Anzahl der Zeilen (Header + 10 Datenzeilen)
        data_rows = 0
        for row in ws.iter_rows(min_row=4, max_row=20, min_col=1, max_col=1):
            if row[0].value is not None:
                data_rows += 1
        
        assert data_rows == 10
    
    def test_export_score_zero(self, tmp_path):
        """Test: Score 0 wird exportiert (neutrales Verhalten ist gültig)"""
        dog_data = DogData(
            owner_name="Test Owner",
            dog_name="Test Dog",
            age_years=2,
            age_months=0,
            gender=Gender.MALE,
            neutered=False
        )
        
        session = TestSession(dog_data=dog_data, battery_name="Test Battery")
        session.add_result(TestResult(test_number=1, score=0, notes="Neutral"))
        session.add_result(TestResult(test_number=2, score=0, notes=""))
        
        exporter = ExcelExporter()
        filepath = tmp_path / "test_export.xlsx"
        
        exporter.export_to_excel(session, str(filepath))
        
        wb = load_workbook(str(filepath))
        ws = wb["Testergebnisse"]
        
        # Beide Tests mit Score 0 sollten exportiert sein
        assert ws.cell(row=4, column=1).value == 1
        assert ws.cell(row=4, column=4).value == 0
        assert ws.cell(row=5, column=1).value == 2
        assert ws.cell(row=5, column=4).value == 0


class TestExcelExporterWithProfiles:
    """Tests für Excel-Export mit OCEAN-Profilen"""
    
    def test_export_with_ocean_profiles(self, tmp_path):
        """Test: Export mit ideal_profile und owner_profile erstellt OCEAN-Analyse Sheet"""
        dog_data = DogData(
            owner_name="Anna Schmidt",
            dog_name="Luna",
            age_years=2,
            age_months=0,
            gender=Gender.FEMALE,
            neutered=True
        )
        
        session = TestSession(dog_data=dog_data, battery_name="OCEAN Battery")
        
        # OCEAN-Profile setzen
        session.ideal_profile = {
            "O": 10, "C": 8, "E": 12, "A": 14, "N": -6
        }
        session.owner_profile = {
            "O": 8, "C": 10, "E": 10, "A": 12, "N": -4
        }
        
        exporter = ExcelExporter()
        filepath = tmp_path / "export_profiles.xlsx"
        
        exporter.export_to_excel(session, str(filepath))
        
        # Excel-Datei laden und OCEAN-Analyse Sheet prüfen
        wb = load_workbook(str(filepath))
        assert "OCEAN-Analyse" in wb.sheetnames
        
        ws = wb["OCEAN-Analyse"]
        
        # Titel prüfen
        assert ws.cell(row=1, column=1).value == "OCEAN-Persönlichkeitsanalyse"
        
        # Header prüfen (Row 3)
        assert ws.cell(row=3, column=1).value == "Dimension"
        assert ws.cell(row=3, column=2).value == "Ist-Profil"
        assert ws.cell(row=3, column=3).value == "Fragebogen-Profil"
        assert ws.cell(row=3, column=4).value == "Ideal-Profil"
        
        # Dimensionen prüfen (Row 4-8)
        dimensions = ["Offenheit (O)", "Gewissenhaftigkeit (C)", "Extraversion (E)", "Verträglichkeit (A)", "Neurotizismus (N)"]
        for i, dim in enumerate(dimensions, start=4):
            assert ws.cell(row=i, column=1).value == dim
        
        # Profile-Werte prüfen (Row 4 = Offenheit)
        assert ws.cell(row=4, column=3).value == 8  # owner_profile O
        assert ws.cell(row=4, column=4).value == 10  # ideal_profile O
        
        # Row 7 = Verträglichkeit
        assert ws.cell(row=7, column=3).value == 12  # owner_profile A
        assert ws.cell(row=7, column=4).value == 14  # ideal_profile A
    
    def test_export_with_ai_assessment(self, tmp_path):
        """Test: Export mit ai_assessment erstellt KI-Bewertung Sheet"""
        dog_data = DogData(
            owner_name="Peter Müller",
            dog_name="Rex",
            age_years=5,
            age_months=3,
            gender=Gender.MALE,
            neutered=False
        )
        
        session = TestSession(dog_data=dog_data, battery_name="Test Battery")
        session.ai_assessment = (
            "Rex zeigt ein ausgeglichenes OCEAN-Profil mit hoher Verträglichkeit "
            "und mittlerer Extraversion. Empfehlung: Mehr Sozialkontakte fördern."
        )
        
        exporter = ExcelExporter()
        filepath = tmp_path / "export_assessment.xlsx"
        
        exporter.export_to_excel(session, str(filepath))
        
        # Excel-Datei laden und KI-Bewertung Sheet prüfen
        wb = load_workbook(str(filepath))
        assert "KI-Bewertung" in wb.sheetnames
        
        ws = wb["KI-Bewertung"]
        
        # Titel prüfen
        assert ws.cell(row=1, column=1).value == "KI-Bewertung"
        
        # Assessment-Text prüfen (in merged cell ab A5)
        assessment_text = str(ws.cell(row=5, column=1).value)
        assert "Rex zeigt ein ausgeglichenes OCEAN-Profil" in assessment_text
        assert "Mehr Sozialkontakte fördern" in assessment_text
    
    def test_export_without_profiles(self, tmp_path):
        """Test: Export ohne Profile erstellt nur 2 Basis-Sheets"""
        dog_data = DogData(
            owner_name="Lisa Wagner",
            dog_name="Bella",
            age_years=1,
            age_months=6,
            gender=Gender.FEMALE,
            neutered=True
        )
        
        session = TestSession(dog_data=dog_data, battery_name="Basic Battery")
        # Keine Profile, keine AI-Bewertung
        
        exporter = ExcelExporter()
        filepath = tmp_path / "export_basic.xlsx"
        
        exporter.export_to_excel(session, str(filepath))
        
        # Excel-Datei laden
        wb = load_workbook(str(filepath))
        
        # Nur Basis-Sheets vorhanden
        assert "Stammdaten" in wb.sheetnames
        assert "Testergebnisse" in wb.sheetnames
        assert "OCEAN-Analyse" not in wb.sheetnames
        assert "KI-Bewertung" not in wb.sheetnames
    
    def test_export_with_battery_calculates_ist_profile(self, tmp_path, sample_battery):
        """Test: Export mit TestBattery berechnet Ist-Profil dynamisch"""
        dog_data = DogData(
            owner_name="Tom Fischer",
            dog_name="Max",
            age_years=4,
            age_months=2,
            gender=Gender.MALE,
            neutered=True
        )
        
        session = TestSession(dog_data=dog_data, battery_name="OCEAN Battery")
        
        # Testergebnisse für alle 3 Tests in battery
        session.add_result(TestResult(test_number=1, score=2, notes=""))  # Openness
        session.add_result(TestResult(test_number=2, score=-1, notes=""))  # Agreeableness
        session.add_result(TestResult(test_number=5, score=1, notes=""))  # Extraversion
        
        # Ideal-Profil setzen
        session.ideal_profile = {
            "O": 10, "C": 5, "E": 8, "A": 12, "N": -3
        }
        
        exporter = ExcelExporter(sample_battery)
        filepath = tmp_path / "export_ist_profile.xlsx"
        
        exporter.export_to_excel(session, str(filepath))
        
        # Excel-Datei laden
        wb = load_workbook(str(filepath))
        ws = wb["OCEAN-Analyse"]
        
        # Ist-Profil prüfen (Spalte B, Rows 4-8)
        # Row 4 = Offenheit (Test 1, score=2)
        assert ws.cell(row=4, column=2).value == 2
        
        # Row 7 = Verträglichkeit (Test 2, score=-1)
        assert ws.cell(row=7, column=2).value == -1
        
        # Row 6 = Extraversion (Test 5, score=1)
        assert ws.cell(row=6, column=2).value == 1
